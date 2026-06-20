import os
from datetime import datetime


RISK_ITEMS = {
    "立杆承载力": {
        "display": "立杆承载力",
        "priority_suggestions": [
            "优先减小立杆纵距、横距（最有效）",
            "减小步距以提高稳定系数",
            "换用壁厚更大的钢管(如φ48×3.5)",
        ],
    },
    "扣件抗滑": {
        "display": "扣件抗滑",
        "priority_suggestions": [
            "优先采用双扣件抗滑（成本最低）",
            "减小立杆间距降低轴力",
            "检查扣件拧紧力矩是否达标",
        ],
    },
    "次楞挠度": {
        "display": "次楞挠度",
        "priority_suggestions": [
            "优先减小次楞间距（见效快）",
            "减小主楞跨度或增加主楞道数",
            "换用更大截面木方(如50×100)",
        ],
    },
    "主楞强度": {
        "display": "主楞强度",
        "priority_suggestions": [
            "优先减小立杆横距（减小主楞跨度）",
            "换用更大规格主楞(如12#槽钢→14#槽钢)",
            "增加立杆以减小主楞计算跨度",
        ],
    },
}


def format_single_result(member_result, warnings=None):
    lines = []
    lines.append("=" * 70)
    lines.append("          模 板 支 撑 验 算 报 告")
    lines.append("=" * 70)

    params = member_result["参数"]
    lines.append(f"工程名称: {params.get('工程名称', '')}")
    lines.append(f"构件名称: {params.get('构件名称', '')}")
    lines.append(f"构件类型: {params.get('构件类型', '')}")
    if params.get("计算人"):
        lines.append(f"计算人:   {params.get('计算人', '')}")
    if params.get("计算日期"):
        lines.append(f"计算日期: {params.get('计算日期', '')}")
    lines.append("-" * 70)

    lines.append("")
    lines.append("【基本参数】")
    lines.append(f"  支撑高度: {params.get('支撑高度', 0):.2f} m")

    if params["构件类型"] == "楼板":
        lines.append(f"  混凝土厚度: {params.get('混凝土厚度', 0):.0f} mm")
    else:
        lines.append(
            f"  梁截面: {params.get('梁宽', 0):.0f} × {params.get('梁高', 0):.0f} mm"
        )

    lines.append(
        f"  立杆纵距 × 横距: "
        f"{params.get('立杆纵距', 0):.2f}m × {params.get('立杆横距', 0):.2f}m"
    )
    lines.append(f"  步距: {params.get('步距', 0):.2f} m")
    lines.append(f"  次楞间距: {params.get('次楞间距', 0):.0f} mm")
    lines.append(f"  主楞间距: {params.get('主楞间距', 0):.2f} m")
    lines.append(f"  立杆钢管: {params.get('立杆钢管规格', '')}")
    lines.append(f"  次楞木方: {params.get('次楞木方规格', '')}")
    lines.append(f"  主楞类型: {params.get('主楞类型', '')}")
    lines.append(f"  扣件类型: {params.get('扣件类型', '')}")
    lines.append(f"  施工活荷载: {params.get('施工活荷载', 0):.1f} kN/m2")

    lines.append("")
    lines.append("-" * 70)
    lines.append("【验算结果汇总】")
    lines.append("-" * 70)

    all_passed = member_result["全部满足"]
    status_str = "全部满足" if all_passed else "存在不满足项"
    lines.append(f"  总体结论: {status_str}")
    lines.append("")

    for name, result in member_result["各项结果"].items():
        status = "[OK] 满足" if result.passed else "[FAIL] 不满足"
        lines.append(f"  {name}: {status}  (比值: {result.ratio:.3f})")

    if not all_passed:
        lines.append("")
        lines.append("  建议调整优先顺序:")
        failed_items = [n for n, r in member_result["各项结果"].items() if not r.passed]
        for i, item in enumerate(failed_items, 1):
            if item in RISK_ITEMS:
                top_sug = RISK_ITEMS[item]["priority_suggestions"][0]
                lines.append(f"    {i}. {RISK_ITEMS[item]['display']}: {top_sug}")

    lines.append("")
    lines.append("=" * 70)
    lines.append("【详细验算过程】")
    lines.append("=" * 70)

    for name, result in member_result["各项结果"].items():
        lines.append("")
        lines.append(f"--- {name} ---")
        lines.append(f"  结论: {'满足要求' if result.passed else '不满足要求'}")
        lines.append(f"  计算值: {result.calculated_value}")
        lines.append(f"  限值:   {result.limit_value}")
        lines.append(f"  比值:   {result.ratio:.3f}")
        lines.append("")
        lines.append("  计算过程:")
        for key, value in result.details.items():
            lines.append(f"    {key}: {value}")
        lines.append("")
        lines.append("  建议:")
        for idx, sug in enumerate(result.suggestions, 1):
            lines.append(f"    {idx}. {sug}")

    if warnings:
        lines.append("")
        lines.append("=" * 70)
        lines.append("【警告信息】")
        lines.append("=" * 70)
        for w in warnings:
            lines.append(f"  [!] {w}")

    lines.append("")
    lines.append("=" * 70)
    lines.append(f"报告生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append("=" * 70)

    return "\n".join(lines)


def _collect_risk_summary(member_results):
    risk_map = {}
    for result in member_results:
        for item_name, check_result in result["各项结果"].items():
            if not check_result.passed:
                if item_name not in risk_map:
                    risk_map[item_name] = []
                risk_map[item_name].append({
                    "构件名称": result["构件名称"],
                    "构件类型": result["构件类型"],
                    "比值": round(check_result.ratio, 3),
                    "建议": check_result.suggestions,
                })

    return risk_map


def format_summary_section(member_results):
    lines = []

    risk_map = _collect_risk_summary(member_results)
    failed_members = [r for r in member_results if not r["全部满足"]]
    passed_members = [r for r in member_results if r["全部满足"]]
    total = len(member_results)

    lines.append("=" * 70)
    lines.append("【工程汇总结论】")
    lines.append("=" * 70)

    lines.append(f"  共验算 {total} 个构件，其中 {len(passed_members)} 个全部满足，"
                 f"{len(failed_members)} 个存在不满足项")

    if not failed_members:
        lines.append("  结论: 全部构件验算通过，可按设计布置施工")
        return "\n".join(lines), risk_map

    lines.append("  结论: 部分构件需调整后再施工")
    lines.append("")
    lines.append("  不满足构件按风险项归类:")
    lines.append("  " + "-" * 60)

    sorted_risks = sorted(
        risk_map.items(),
        key=lambda x: len(x[1]),
        reverse=True
    )

    for item_name, failed_list in sorted_risks:
        count = len(failed_list)
        display = RISK_ITEMS.get(item_name, {}).get("display", item_name)
        lines.append(f"")
        lines.append(f"  [{display}] 共 {count} 个构件不满足:")

        for i, info in enumerate(failed_list, 1):
            lines.append(f"    {i}. {info['构件名称']} ({info['构件类型']})"
                         f" - 比值 {info['比值']}")

        lines.append(f"  >> 优先调整方向:")
        if item_name in RISK_ITEMS:
            for j, sug in enumerate(RISK_ITEMS[item_name]["priority_suggestions"], 1):
                lines.append(f"     {j}. {sug}")
        else:
            for sug in failed_list[0]["建议"][:2]:
                lines.append(f"     - {sug}")

    lines.append("")
    lines.append("  >> 整体复核建议:")
    if "立杆承载力" in risk_map and "扣件抗滑" in risk_map:
        lines.append(
            "     立杆与扣件同时存在不满足，建议优先检查立杆间距是否过大，"
            "必要时加密立杆布置"
        )
    if "次楞挠度" in risk_map and "主楞强度" in risk_map:
        lines.append(
            "     次楞与主楞同时存在不满足，建议复核模板体系的龙骨布置，"
            "考虑减小主楞跨度或增加龙骨截面"
        )
    if len(risk_map) >= 3:
        lines.append(
            "     多项指标同时不满足，建议整体降低立杆间距、"
            "步距，并加大龙骨规格后重新验算"
        )

    return "\n".join(lines), risk_map


def format_batch_result(batch_data, member_results, warnings=None, row_errors=None):
    lines = []
    lines.append("=" * 70)
    lines.append("      模 板 支 撑 批 量 验 算 报 告")
    lines.append("=" * 70)

    lines.append(f"工程名称: {batch_data.get('工程名称', '')}")
    if batch_data.get("计算人"):
        lines.append(f"计算人:   {batch_data.get('计算人', '')}")
    if batch_data.get("计算日期"):
        lines.append(f"计算日期: {batch_data.get('计算日期', '')}")
    lines.append(f"构件数量: {len(member_results)} 个")
    if row_errors:
        lines.append(f"跳过行数: {len(row_errors)} 行(参数有误)")
    lines.append("-" * 70)

    if row_errors:
        lines.append("")
        lines.append("=" * 70)
        lines.append("【参数错误行清单】")
        lines.append("=" * 70)
        lines.append(f"  以下 {len(row_errors)} 行因参数错误未参与验算，请修正后重新运行:")
        for err in row_errors:
            lines.append(f"  - 第{err['行号']}行 [{err.get('构件名', '')}]: {err['错误']}")

    summary_text, risk_map = format_summary_section(member_results)
    lines.append("")
    lines.append(summary_text)

    lines.append("")
    lines.append("-" * 70)
    lines.append("【验算结果汇总表】")
    lines.append("-" * 70)

    header = f"{'序号':<4}{'构件名称':<25}{'类型':<6}"
    check_items = list(member_results[0]["各项结果"].keys()) if member_results else []
    for item in check_items:
        header += f"{item[:4]:<8}"
    header += " 结论"
    lines.append(header)
    lines.append("-" * 70)

    passed_count = 0
    for idx, result in enumerate(member_results, 1):
        row = f"{idx:<4}{result['构件名称'][:24]:<25}{result['构件类型']:<6}"
        for item in check_items:
            r = result["各项结果"][item]
            status = "OK" if r.passed else "FAIL"
            row += f"{status:<8}"
        overall = "通过" if result["全部满足"] else "不通过"
        row += f" {overall}"
        lines.append(row)
        if result["全部满足"]:
            passed_count += 1

    lines.append("-" * 70)
    lines.append(
        f"汇总: {passed_count}/{len(member_results)} 个构件全部满足要求"
    )

    failed_members = [r for r in member_results if not r["全部满足"]]
    if failed_members:
        lines.append("")
        lines.append("=" * 70)
        lines.append("【不满足项逐构件详情】")
        lines.append("=" * 70)

        for result in failed_members:
            lines.append("")
            lines.append(f"* {result['构件名称']} ({result['构件类型']})")
            for name, check_result in result["各项结果"].items():
                if not check_result.passed:
                    lines.append(f"  > {name}: 不满足 (比值: {check_result.ratio:.3f})")
                    for sug in check_result.suggestions:
                        lines.append(f"     建议: {sug}")

    lines.append("")
    lines.append("=" * 70)
    lines.append("【各构件验算概要】")
    lines.append("=" * 70)

    for idx, result in enumerate(member_results, 1):
        lines.append("")
        lines.append(f"--- [{idx}] {result['构件名称']} ---")
        params = result["参数"]
        if params["构件类型"] == "楼板":
            lines.append(
                f"  板厚{params.get('混凝土厚度', 0):.0f}mm, "
                f"立杆{params.get('立杆纵距', 0):.2f}×"
                f"{params.get('立杆横距', 0):.2f}m, "
                f"步距{params.get('步距', 0):.2f}m, "
                f"高{params.get('支撑高度', 0):.2f}m"
            )
        else:
            lines.append(
                f"  梁{params.get('梁宽', 0):.0f}×{params.get('梁高', 0):.0f}mm, "
                f"立杆{params.get('立杆纵距', 0):.2f}×"
                f"{params.get('立杆横距', 0):.2f}m, "
                f"步距{params.get('步距', 0):.2f}m, "
                f"高{params.get('支撑高度', 0):.2f}m"
            )

        for name, check_result in result["各项结果"].items():
            status = "满足" if check_result.passed else "不满足"
            lines.append(
                f"  {name}: {status} "
                f"(计算值={check_result.calculated_value}, "
                f"限值={check_result.limit_value}, "
                f"比值={check_result.ratio:.3f})"
            )

    if warnings:
        lines.append("")
        lines.append("=" * 70)
        lines.append("【警告信息】")
        lines.append("=" * 70)
        for w in warnings:
            lines.append(f"  [!] {w}")

    lines.append("")
    lines.append("=" * 70)
    lines.append(f"报告生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append("=" * 70)

    return "\n".join(lines)


def generate_batch_with_errors(batch_data, member_results, warnings=None, row_errors=None):
    return format_batch_result(batch_data, member_results, warnings, row_errors)


def save_report(content, output_path):
    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else ".", exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(content)
    return output_path


def generate_output(parsed_data, results, warnings=None, output_path=None):
    if parsed_data["模式"] == "single":
        content = format_single_result(results, warnings)
    else:
        content = format_batch_result(parsed_data, results, warnings)

    if output_path:
        save_report(content, output_path)

    return content


CHECK_ITEMS = ["立杆承载力", "扣件抗滑", "次楞挠度", "主楞强度"]


def _build_export_rows(member_results, row_errors=None):
    """
    构建结构化导出的行数据
    返回: (result_rows, error_rows)
        result_rows: 每个构件一行
        error_rows: 每个参数错误一行
    """
    result_rows = []
    for idx, mr in enumerate(member_results, 1):
        params = mr.get("参数", {})
        member_name = mr.get("构件名称", "")
        member_type = mr.get("构件类型", "")
        all_pass = mr.get("全部满足", False)
        results_dict = mr.get("各项结果", {})

        row = {
            "序号": idx,
            "构件名称": member_name,
            "构件类型": member_type,
            "总体结论": "通过" if all_pass else "不通过",
        }

        fail_items = []
        all_suggestions = []
        for item in CHECK_ITEMS:
            if item in results_dict:
                r = results_dict[item]
                passed = r.passed
                ratio = round(r.ratio, 3)
                status = "OK" if passed else "FAIL"
                row[f"{item}_状态"] = status
                row[f"{item}_比值"] = ratio
                if not passed:
                    fail_items.append(item)
                    if r.suggestions:
                        all_suggestions.extend(r.suggestions[:2])

        row["不满足项"] = "; ".join(fail_items)
        row["主要建议"] = "; ".join(all_suggestions[:3])

        row["立杆纵距(m)"] = params.get("立杆纵距", "")
        row["立杆横距(m)"] = params.get("立杆横距", "")
        row["步距(m)"] = params.get("步距", "")
        row["次楞间距(mm)"] = params.get("次楞间距", "")
        row["主楞间距(m)"] = params.get("主楞间距", "")
        row["混凝土厚度(mm)"] = params.get("混凝土厚度", params.get("梁宽", ""))
        if member_type == "梁":
            row["梁宽(mm)"] = params.get("梁宽", "")
            row["梁高(mm)"] = params.get("梁高", "")
        else:
            row["支撑高度(m)"] = params.get("支撑高度", "")
        row["立杆钢管"] = params.get("立杆钢管规格", "")
        row["次楞木方"] = params.get("次楞木方规格", "")
        row["主楞类型"] = params.get("主楞类型", "")
        row["扣件类型"] = params.get("扣件类型", "")
        row["施工活荷载(kN/m2)"] = params.get("施工活荷载", "")
        result_rows.append(row)

    error_rows = []
    if row_errors:
        for err in row_errors:
            error_rows.append({
                "行号": err.get("行号", ""),
                "构件名": err.get("构件名", ""),
                "错误信息": err.get("错误", ""),
            })

    return result_rows, error_rows


CSV_RESULT_COLUMNS = [
    "序号", "构件名称", "构件类型", "总体结论",
    "立杆承载力_状态", "立杆承载力_比值",
    "扣件抗滑_状态", "扣件抗滑_比值",
    "次楞挠度_状态", "次楞挠度_比值",
    "主楞强度_状态", "主楞强度_比值",
    "不满足项", "主要建议",
    "立杆纵距(m)", "立杆横距(m)", "步距(m)", "次楞间距(mm)", "主楞间距(m)",
    "混凝土厚度(mm)", "梁宽(mm)", "梁高(mm)", "支撑高度(m)",
    "立杆钢管", "次楞木方", "主楞类型", "扣件类型", "施工活荷载(kN/m2)"
]

CSV_ERROR_COLUMNS = ["行号", "构件名", "错误信息"]


def export_structured_results(member_results, output_path, file_format="csv", row_errors=None, project_info=None):
    """
    导出结构化结果表到 CSV 或 Excel
    file_format: "csv" 或 "xlsx"
    """
    file_format = file_format.lower().lstrip(".")
    result_rows, error_rows = _build_export_rows(member_results, row_errors)

    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else ".", exist_ok=True)

    if file_format in ("csv",):
        with open(output_path, "w", encoding="utf-8-sig", newline="") as f:
            import csv as _csv
            writer = _csv.writer(f)
            if project_info:
                writer.writerow(["工程名称", project_info.get("工程名称", "")])
                writer.writerow(["计算人", project_info.get("计算人", "")])
                writer.writerow(["计算日期", project_info.get("计算日期", "")])
                writer.writerow([])

            writer.writerow(["==== 验算结果"])
            writer.writerow(CSV_RESULT_COLUMNS)
            for row in result_rows:
                writer.writerow([row.get(col, "") for col in CSV_RESULT_COLUMNS])

            if error_rows:
                writer.writerow([])
                writer.writerow(["==== 参数错误行"])
                writer.writerow(CSV_ERROR_COLUMNS)
                for err in error_rows:
                    writer.writerow([err.get(col, "") for col in CSV_ERROR_COLUMNS])

        return output_path

    elif file_format in ("xlsx", "xls"):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise RuntimeError("未安装 openpyxl，无法导出 Excel，请先安装: pip install openpyxl")

        wb = Workbook()

        ws_result = wb.active
        ws_result.title = "验算结果"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        center = Alignment(horizontal="center", vertical="center")

        current_row = 1
        if project_info:
            ws_result.cell(row=current_row, column=1, value="工程名称").font = Font(bold=True)
            ws_result.cell(row=current_row, column=2, value=project_info.get("工程名称", ""))
            current_row += 1
            ws_result.cell(row=current_row, column=1, value="计算人").font = Font(bold=True)
            ws_result.cell(row=current_row, column=2, value=project_info.get("计算人", ""))
            current_row += 1
            ws_result.cell(row=current_row, column=1, value="计算日期").font = Font(bold=True)
            ws_result.cell(row=current_row, column=2, value=project_info.get("计算日期", ""))
            current_row += 2

        header_row_idx = current_row
        for col_idx, col_name in enumerate(CSV_RESULT_COLUMNS, 1):
            cell = ws_result.cell(row=header_row_idx, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
        current_row += 1

        for row_data in result_rows:
            overall = row_data.get("总体结论", "")
            for col_idx, col_name in enumerate(CSV_RESULT_COLUMNS, 1):
                val = row_data.get(col_name, "")
                cell = ws_result.cell(row=current_row, column=col_idx, value=val)
                cell.alignment = center
                if overall == "不通过":
                    cell.fill = fail_fill
                elif col_name.endswith("_状态") and val == "OK":
                    cell.fill = pass_fill
            current_row += 1

        for col_idx in range(1, len(CSV_RESULT_COLUMNS) + 1):
            ws_result.column_dimensions[get_column_letter(col_idx)].width = max(12, len(CSV_RESULT_COLUMNS[col_idx - 1]) + 2)

        if error_rows:
            ws_err = wb.create_sheet("参数错误行")
            for col_idx, col_name in enumerate(CSV_ERROR_COLUMNS, 1):
                cell = ws_err.cell(row=1, column=col_idx, value=col_name)
                cell.font = header_font
                cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
                cell.alignment = center
            for r_idx, err_data in enumerate(error_rows, 2):
                for col_idx, col_name in enumerate(CSV_ERROR_COLUMNS, 1):
                    cell = ws_err.cell(row=r_idx, column=col_idx, value=err_data.get(col_name, ""))
                    cell.alignment = center
            for col_idx in range(1, len(CSV_ERROR_COLUMNS) + 1):
                ws_err.column_dimensions[get_column_letter(col_idx)].width = max(12, len(CSV_ERROR_COLUMNS[col_idx - 1]) + 2)

        wb.save(output_path)
        return output_path

    else:
        raise ValueError(f"不支持的导出格式: {file_format}，请使用 csv 或 xlsx")
